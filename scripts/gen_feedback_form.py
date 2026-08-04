"""
Generate expert_feedback_form.xlsx for chatbot user testing.

Usage:
    python scripts/gen_feedback_form.py

Output:
    data/expert_feedback_form.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.protection import Protection

# ── Color palette ──────────────────────────────────────────────────────────────
NAVY       = "1e3a5f"
BLUE_LIGHT = "dbeafe"
YELLOW_IN  = "fffde7"   # input cell highlight
WHITE      = "ffffff"
GRAY_LIGHT = "f8fafc"
RED_LIGHT  = "fee2e2"
GREEN_LIGHT = "dcfce7"
HEADER_FG  = "ffffff"

# ── Style helpers ──────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="d1d5db")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(bold=True, size=11, color=HEADER_FG):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def body_font(bold=False, size=11, color="1e293b"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_wrap():
    return Alignment(horizontal="left", vertical="top", wrap_text=True)

def apply_header(cell, text, bg=NAVY, fg=HEADER_FG, size=11, bold=True):
    cell.value = text
    cell.fill = fill(bg)
    cell.font = Font(name="Calibri", bold=bold, size=size, color=fg)
    cell.alignment = center()
    cell.border = thin_border()

def apply_input(cell, value=""):
    cell.value = value
    cell.fill = fill(YELLOW_IN)
    cell.font = body_font()
    cell.alignment = left_wrap()
    cell.border = thin_border()

def apply_label(cell, text, bold=False, bg=WHITE):
    cell.value = text
    cell.fill = fill(bg)
    cell.font = body_font(bold=bold)
    cell.alignment = left_wrap()
    cell.border = thin_border()

def apply_formula(cell, formula, bg=GRAY_LIGHT):
    cell.value = formula
    cell.fill = fill(bg)
    cell.font = body_font(color="374151")
    cell.alignment = center()
    cell.border = thin_border()

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height


# ── Sheet 1: Profil & Panduan ──────────────────────────────────────────────────

def build_sheet1(wb):
    ws = wb.create_sheet("Profil & Panduan")
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "FORMULIR EVALUASI EXPERT — K8s GraphRAG Chatbot"
    c.fill = fill(NAVY)
    c.font = Font(name="Calibri", bold=True, size=14, color=HEADER_FG)
    c.alignment = center()
    set_row_height(ws, 1, 36)

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "Jihan Aurelia · Tugas Akhir STI ITB · 2026"
    c.fill = fill("2563eb")
    c.font = Font(name="Calibri", size=11, color=HEADER_FG)
    c.alignment = center()
    set_row_height(ws, 2, 22)

    # Link chatbot
    ws.merge_cells("A3:F3")
    c = ws["A3"]
    c.value = "Akses Chatbot: https://huggingface.co/spaces/jijiau/k8s-graphrag-chatbot"
    c.fill = fill(BLUE_LIGHT)
    c.font = Font(name="Calibri", size=11, color="1e40af", underline="single")
    c.alignment = center()
    ws.row_dimensions[3].height = 20

    ws.row_dimensions[4].height = 8  # spacer

    # Identitas Expert
    ws.merge_cells("A5:F5")
    apply_header(ws["A5"], "IDENTITAS EXPERT", bg="374151", size=11)
    set_row_height(ws, 5, 22)

    labels_identitas = [
        ("Nama Lengkap", ""),
        ("Institusi / Perusahaan", ""),
        ("Jabatan / Posisi", ""),
        ("Tahun Pengalaman dengan Kubernetes", ""),
        ("Resource K8s yang Paling Sering Digunakan", "contoh: Deployment, Service, Ingress, ..."),
    ]
    for i, (label, hint) in enumerate(labels_identitas, start=6):
        ws.merge_cells(f"A{i}:B{i}")
        apply_label(ws[f"A{i}"], label, bold=True, bg=GRAY_LIGHT)
        ws.merge_cells(f"C{i}:F{i}")
        apply_input(ws[f"C{i}"], hint)
        set_row_height(ws, i, 22)

    ws.row_dimensions[11].height = 8  # spacer

    # Panduan pengisian
    ws.merge_cells("A12:F12")
    apply_header(ws["A12"], "PANDUAN PENGISIAN", bg="374151", size=11)
    set_row_height(ws, 12, 22)

    panduan = [
        "1.  Buka link chatbot di atas menggunakan browser. Pastikan Anda sudah mendapat akses dari Jihan.",
        "2.  Sheet \"Skenario Uji\" berisi 8 pertanyaan panduan. Skenario ini OPSIONAL — Anda bebas menggantinya\n     dengan pertanyaan K8s dari pengalaman kerja Anda sendiri.",
        "3.  Untuk setiap pertanyaan yang dicoba, salin respons chatbot ke kolom \"Respons Chatbot\", lalu beri\n     nilai Akurasi dan Kejelasan menggunakan dropdown (1=Sangat Buruk, 5=Sangat Baik).",
        "4.  Catat bug atau error yang ditemukan di sheet \"Temuan & Saran\". Isi juga kolom Saran & Kesan Umum.",
        "5.  Sheet \"Rekap Otomatis\" terisi sendiri — tidak perlu diisi manual.",
        "6.  Kirimkan file yang sudah diisi ke: jihanaurelia.jiji@gmail.com",
    ]
    for i, text in enumerate(panduan, start=13):
        ws.merge_cells(f"A{i}:F{i}")
        c = ws[f"A{i}"]
        c.value = text
        c.fill = fill(WHITE)
        c.font = body_font(size=11)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = thin_border()
        set_row_height(ws, i, 32)

    # Column widths
    col_widths = [20, 20, 18, 18, 18, 18]
    for i, w in enumerate(col_widths, 1):
        set_col_width(ws, i, w)

    ws.freeze_panes = "A4"
    return ws


# ── Sheet 2: Skenario Uji ──────────────────────────────────────────────────────

SKENARIO = [
    ("Conceptual",         "Apa perbedaan Deployment dan StatefulSet?"),
    ("Conceptual",         "Kapan harus menggunakan DaemonSet dibandingkan Deployment?"),
    ("Generate YAML",      "Buat YAML Deployment nginx dengan 3 replica dan resource limits (CPU 500m, memory 256Mi)"),
    ("Generate YAML",      "Buat YAML untuk CronJob yang menjalankan backup database setiap hari pukul 02.00"),
    ("Trace Relationship", "Bagaimana Service terhubung ke Pod di Kubernetes?"),
    ("Trace Relationship", "Apa hubungan antara HorizontalPodAutoscaler (HPA) dan Deployment?"),
    ("Planning",           "Saya perlu isolasi antara tim frontend dan backend di cluster yang sama — apa saja yang perlu dikonfigurasi?"),
    ("Planning",           "Aplikasi saya perlu bisa scale otomatis saat traffic tinggi dan tetap tersedia tanpa downtime saat di-update — setup apa yang dibutuhkan?"),
]

def build_sheet2(wb):
    ws = wb.create_sheet("Skenario Uji")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "SKENARIO UJI — Chatbot K8s GraphRAG"
    c.fill = fill(NAVY)
    c.font = Font(name="Calibri", bold=True, size=13, color=HEADER_FG)
    c.alignment = center()
    set_row_height(ws, 1, 32)

    # Keterangan opsional
    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = (
        "Skenario berikut bersifat OPSIONAL sebagai panduan awal. "
        "Anda bebas mengganti atau menambah pertanyaan sesuai pengalaman K8s Anda sendiri."
    )
    c.fill = fill(GREEN_LIGHT)
    c.font = Font(name="Calibri", size=10, italic=True, color="166534")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = thin_border()
    set_row_height(ws, 2, 30)

    ws.row_dimensions[3].height = 6  # spacer

    # Column headers
    headers = ["#", "Kategori", "Pertanyaan yang Dicoba", "Respons Chatbot",
               "Akurasi\n(1-5)", "Kejelasan\n(1-5)", "YAML Valid?", "Catatan"]
    bg_row = ["374151"] * len(headers)
    for col, (h, bg) in enumerate(zip(headers, bg_row), 1):
        apply_header(ws.cell(4, col), h, bg="374151")
    set_row_height(ws, 4, 36)

    # Data validation dropdowns
    dv_15 = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True,
                           showInputMessage=True, showErrorMessage=True)
    dv_15.prompt = "1=Sangat Buruk, 5=Sangat Baik"
    dv_15.error = "Pilih nilai 1-5"
    dv_15.sqref = f"E5:F12"
    ws.add_data_validation(dv_15)

    dv_yaml = DataValidation(type="list", formula1='"Ya,Tidak,Tidak Diuji"', allow_blank=True)
    dv_yaml.sqref = "G5:G12"
    ws.add_data_validation(dv_yaml)

    # Skenario rows
    for i, (kategori, pertanyaan) in enumerate(SKENARIO, start=5):
        row = i
        is_yaml = "YAML" in kategori

        # Alternating row color
        row_bg = BLUE_LIGHT if (i % 2 == 0) else WHITE

        # #
        c = ws.cell(row, 1)
        c.value = i - 4
        c.fill = fill(row_bg)
        c.font = body_font(bold=True)
        c.alignment = center()
        c.border = thin_border()

        # Kategori
        c = ws.cell(row, 2)
        c.value = kategori
        c.fill = fill(row_bg)
        c.font = body_font(bold=False, color="1e40af" if "YAML" in kategori else
                           "065f46" if "Planning" in kategori else
                           "92400e" if "Trace" in kategori else "374151")
        c.alignment = center()
        c.border = thin_border()

        # Pertanyaan (pre-filled, italic)
        c = ws.cell(row, 3)
        c.value = pertanyaan
        c.fill = fill(row_bg)
        c.font = Font(name="Calibri", italic=True, size=11, color="374151")
        c.alignment = left_wrap()
        c.border = thin_border()

        # Respons Chatbot (input)
        apply_input(ws.cell(row, 4))

        # Akurasi (input + dropdown)
        apply_input(ws.cell(row, 5))
        ws.cell(row, 5).alignment = center()

        # Kejelasan (input + dropdown)
        apply_input(ws.cell(row, 6))
        ws.cell(row, 6).alignment = center()

        # YAML Valid? — only yellow for YAML rows
        c = ws.cell(row, 7)
        if is_yaml:
            apply_input(c)
        else:
            c.value = "N/A"
            c.fill = fill(GRAY_LIGHT)
            c.font = body_font(color="9ca3af")
            c.alignment = center()
            c.border = thin_border()

        # Catatan (input)
        apply_input(ws.cell(row, 8))

        set_row_height(ws, row, 60)

    # Extra empty rows for expert's own questions
    for i in range(13, 16):
        for col in range(1, 9):
            c = ws.cell(i, col)
            if col == 1:
                c.value = i - 4
                c.fill = fill(WHITE)
                c.font = body_font(bold=True)
                c.alignment = center()
                c.border = thin_border()
            elif col == 2:
                apply_input(c)
            else:
                apply_input(c)
                if col in (5, 6):
                    c.alignment = center()
        set_row_height(ws, i, 60)

    # Column widths
    col_widths = [4, 14, 38, 38, 9, 9, 11, 24]
    for i, w in enumerate(col_widths, 1):
        set_col_width(ws, i, w)

    ws.freeze_panes = "A5"
    return ws


# ── Sheet 3: Temuan & Saran ────────────────────────────────────────────────────

def build_sheet3(wb):
    ws = wb.create_sheet("Temuan & Saran")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "TEMUAN & SARAN"
    c.fill = fill(NAVY)
    c.font = Font(name="Calibri", bold=True, size=13, color=HEADER_FG)
    c.alignment = center()
    set_row_height(ws, 1, 32)

    ws.row_dimensions[2].height = 6

    # Bug table header
    ws.merge_cells("A3:E3")
    apply_header(ws["A3"], "Bug / Error yang Ditemukan", bg="374151")
    set_row_height(ws, 3, 22)

    bug_headers = ["#", "Deskripsi Masalah", "Langkah Reproduce", "Tingkat Keparahan", "Catatan"]
    for col, h in enumerate(bug_headers, 1):
        apply_header(ws.cell(4, col), h, bg="4b5563")
    set_row_height(ws, 4, 28)

    dv_sev = DataValidation(type="list", formula1='"Low,Medium,High"', allow_blank=True)
    dv_sev.sqref = "D5:D14"
    ws.add_data_validation(dv_sev)

    for i in range(5, 15):
        row_bg = BLUE_LIGHT if i % 2 == 0 else WHITE
        c = ws.cell(i, 1)
        c.value = i - 4
        c.fill = fill(row_bg)
        c.font = body_font(bold=True)
        c.alignment = center()
        c.border = thin_border()
        for col in range(2, 6):
            c = ws.cell(i, col)
            apply_input(c)
            if col == 4:
                c.alignment = center()
        set_row_height(ws, i, 40)

    ws.row_dimensions[15].height = 10

    # Saran & Kesan Umum
    ws.merge_cells("A16:E16")
    apply_header(ws["A16"], "Saran & Kesan Umum", bg="374151")
    set_row_height(ws, 16, 22)

    ws.merge_cells("A17:E26")
    c = ws["A17"]
    c.value = "Tuliskan saran, pendapat, atau kesan umum Anda di sini..."
    c.fill = fill(YELLOW_IN)
    c.font = Font(name="Calibri", size=11, color="9ca3af", italic=True)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border = thin_border()
    for i in range(17, 27):
        set_row_height(ws, i, 22)

    # Column widths
    col_widths = [4, 32, 32, 16, 20]
    for i, w in enumerate(col_widths, 1):
        set_col_width(ws, i, w)

    ws.freeze_panes = "A5"
    return ws


# ── Sheet 4: Rekap Otomatis ────────────────────────────────────────────────────

def build_sheet4(wb):
    ws = wb.create_sheet("Rekap Otomatis")
    ws.sheet_view.showGridLines = False
    ws.protection.sheet = True
    ws.protection.password = "readonly"

    # Title
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "REKAP OTOMATIS — Tidak Perlu Diisi Manual"
    c.fill = fill(NAVY)
    c.font = Font(name="Calibri", bold=True, size=13, color=HEADER_FG)
    c.alignment = center()
    set_row_height(ws, 1, 32)

    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = "Sheet ini terisi otomatis dari data di Sheet 'Skenario Uji' dan 'Temuan & Saran'."
    c.fill = fill(BLUE_LIGHT)
    c.font = Font(name="Calibri", size=10, italic=True, color="1e40af")
    c.alignment = center()
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8

    # Rekap skenario
    ws.merge_cells("A4:D4")
    apply_header(ws["A4"], "Rata-rata Penilaian per Skenario", bg="374151")
    set_row_height(ws, 4, 22)

    rekap_headers = ["#", "Pertanyaan", "Rata-rata Akurasi", "Rata-rata Kejelasan"]
    for col, h in enumerate(rekap_headers, 1):
        apply_header(ws.cell(5, col), h, bg="4b5563")
    set_row_height(ws, 5, 28)

    for i, (_, pertanyaan) in enumerate(SKENARIO, start=6):
        row_bg = BLUE_LIGHT if i % 2 == 0 else WHITE
        ws.cell(i, 1).value = i - 5
        ws.cell(i, 1).fill = fill(row_bg)
        ws.cell(i, 1).font = body_font(bold=True)
        ws.cell(i, 1).alignment = center()
        ws.cell(i, 1).border = thin_border()

        c = ws.cell(i, 2)
        c.value = pertanyaan[:80] + "..." if len(pertanyaan) > 80 else pertanyaan
        c.fill = fill(row_bg)
        c.font = body_font()
        c.alignment = left_wrap()
        c.border = thin_border()

        skenario_row = i - 5 + 4  # row di sheet Skenario Uji (row 5-12)
        for col, src_col in [(3, "E"), (4, "F")]:
            apply_formula(ws.cell(i, col), f"='Skenario Uji'!{src_col}{skenario_row}", bg=row_bg)
        set_row_height(ws, i, 36)

    # Total rata-rata
    total_row = 6 + len(SKENARIO)
    ws.merge_cells(f"A{total_row}:B{total_row}")
    c = ws[f"A{total_row}"]
    c.value = "TOTAL RATA-RATA"
    c.fill = fill(NAVY)
    c.font = Font(name="Calibri", bold=True, size=11, color=HEADER_FG)
    c.alignment = center()
    c.border = thin_border()

    for col, col_letter in [(3, "C"), (4, "D")]:
        apply_formula(ws.cell(total_row, col),
                      f"=AVERAGE({col_letter}6:{col_letter}{total_row - 1})",
                      bg=NAVY)
        ws.cell(total_row, col).font = Font(name="Calibri", bold=True, size=11, color=HEADER_FG)
    set_row_height(ws, total_row, 28)

    ws.row_dimensions[total_row + 1].height = 10

    # Bug severity summary
    summary_start = total_row + 2
    ws.merge_cells(f"A{summary_start}:D{summary_start}")
    apply_header(ws[f"A{summary_start}"], "Ringkasan Bug per Severity", bg="374151")
    set_row_height(ws, summary_start, 22)

    severity_rows = [
        ("High",   RED_LIGHT,   "High"),
        ("Medium", YELLOW_IN,   "Medium"),
        ("Low",    GREEN_LIGHT, "Low"),
    ]
    for j, (label, color, val) in enumerate(severity_rows, start=summary_start + 1):
        ws.merge_cells(f"A{j}:C{j}")
        c = ws[f"A{j}"]
        c.value = f"Jumlah Bug {label}"
        c.fill = fill(color)
        c.font = body_font(bold=True)
        c.alignment = left_wrap()
        c.border = thin_border()
        apply_formula(ws.cell(j, 4),
                      f"=COUNTIF('Temuan & Saran'!D5:D14,\"{val}\")",
                      bg=color)
        set_row_height(ws, j, 22)

    # Column widths
    col_widths = [4, 55, 18, 18]
    for i, w in enumerate(col_widths, 1):
        set_col_width(ws, i, w)

    ws.freeze_panes = "A5"
    return ws


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    build_sheet1(wb)
    build_sheet2(wb)
    build_sheet3(wb)
    build_sheet4(wb)

    output_path = os.path.join(os.path.dirname(__file__), "..", "data", "expert_feedback_form.xlsx")
    output_path = os.path.normpath(output_path)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Generated: {output_path}")


if __name__ == "__main__":
    main()
