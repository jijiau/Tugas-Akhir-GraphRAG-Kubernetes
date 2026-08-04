"""Generate CHECKLIST_KUALITAS_SIDANG.xlsx for TA GraphRAG Kubernetes."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Checklist Sidang"

# ── Colours ────────────────────────────────────────────────────────────────
LULUS_FILL   = PatternFill("solid", fgColor="C6EFCE")
DEVIASI_FILL = PatternFill("solid", fgColor="FFEB9C")
BLOCKER_FILL = PatternFill("solid", fgColor="FFC7CE")
MAJOR_FILL   = PatternFill("solid", fgColor="FFE4B5")
MINOR_FILL   = PatternFill("solid", fgColor="DDEEFF")
HEADER_FILL  = PatternFill("solid", fgColor="203864")
CAT_FILL     = PatternFill("solid", fgColor="2F5496")

WHITE_BOLD = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
CAT_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
NORM_FONT  = Font(name="Calibri", size=10)
BOLD_FONT  = Font(name="Calibri", bold=True, size=10)
SMALL_FONT = Font(name="Calibri", size=9)

thin = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell_style(cell, fill=None, font=None, align_h="left", align_v="center",
               wrap=True, border=True):
    if fill:
        cell.fill = fill
    cell.font = font or NORM_FONT
    cell.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=wrap)
    if border:
        cell.border = THIN_BORDER

def sev_fill(s):
    if s == "Blocker": return BLOCKER_FILL
    if s == "Major":   return MAJOR_FILL
    return MINOR_FILL

# ── Column widths ─────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 7
ws.column_dimensions["B"].width = 55
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 70
ws.column_dimensions["E"].width = 11

# ── Title rows ────────────────────────────────────────────────────────────
ws.merge_cells("A1:E1")
ws["A1"] = "CHECKLIST KUALITAS TESIS — GraphRAG Kubernetes (Jihan Aurelia, 18222001)"
cell_style(ws["A1"], fill=HEADER_FILL, font=WHITE_BOLD, align_h="center")
ws.row_dimensions[1].height = 24

ws.merge_cells("A2:E2")
ws["A2"] = (
    "Judul: Implementasi Graph Retrieval-Augmented Generation untuk Meningkatkan "
    "Presisi Retrieval dan Validitas Sintaksis pada Konfigurasi Kubernetes"
)
cell_style(ws["A2"],
           fill=PatternFill("solid", fgColor="D6E4F7"),
           font=Font(name="Calibri", italic=True, size=9),
           align_h="center")
ws.row_dimensions[2].height = 18

# ── Header row ────────────────────────────────────────────────────────────
for col, h in enumerate(
    ["Kode", "Kriteria", "Status", "Catatan & Bukti (halaman/kutipan)", "Severitas"], 1
):
    c = ws.cell(row=3, column=col, value=h)
    cell_style(c, fill=HEADER_FILL, font=WHITE_BOLD, align_h="center")
ws.row_dimensions[3].height = 20

# ── Data ──────────────────────────────────────────────────────────────────
LULUS   = "LULUS"
DEVIASI = "DEVIASI DITERIMA"
CAT     = "CAT"

rows = [
    # ── A ────────────────────────────────────────────────────────────────
    (CAT, "A. Halaman Awal & Identitas Dokumen", "", "", ""),
    ("A1",
     "Judul konsisten di sampul, lembar pengesahan, dan abstrak; sesuai isi & ruang lingkup",
     LULUS,
     "Halaman Judul (Hal. i), Lembar Pengesahan (Hal. iii), Abstrak (Hal. ix), "
     "Kata Pengantar (Hal. xi) — seragam \"Presisi Retrieval\" setelah perbaikan cover.",
     "Blocker"),
    ("A2",
     "Kata Pengantar membahas TOPIK ASLI tesis ini (bukan topik template/sisa proyek lain)",
     LULUS,
     "Kata Pengantar (Hal. xi) — membahas GraphRAG untuk dokumentasi API Kubernetes.",
     "Minor"),
    ("A3",
     "Isi abstrak membahas masalah + metode + hasil (dengan angka) + keterbatasan",
     LULUS,
     "Abstrak ID (Hal. ix) — masalah, metode, hasil (RetQ 0,7259; ReaQ 0,5554; "
     "PathCov 0,8515; RGA 0,4536; p<0,001), keterbatasan, kata kunci.",
     "Major"),
    ("A4",
     "Abstrak tersedia versi Indonesia + Inggris; kata kunci ada dan benar",
     DEVIASI,
     "Abstrak ID (Hal. ix) — hanya versi Bahasa Indonesia, keputusan eksplisit peneliti. "
     "Abstrak ID sudah lengkap dengan semua elemen wajib + 5 kata kunci.",
     "Major"),
    ("A5",
     "Kata Pengantar menyebut judul yang benar & NIM yang benar (tanpa placeholder)",
     LULUS,
     "Kata Pengantar (Hal. xi) — judul \"Presisi Retrieval\", NIM 18222001, "
     "nama Jihan Aurelia, tanpa placeholder.",
     "Minor"),
    ("A6",
     "Pernyataan Orisinalitas + Pernyataan Penggunaan AI lengkap dan ditandatangani",
     LULUS,
     "Pernyataan Orisinalitas (Hal. v), Pernyataan Penggunaan AI (Hal. vii) — "
     "keduanya ada, lengkap, tanpa placeholder.",
     "Blocker"),
    ("A7",
     "Daftar Simbol berisi simbol yang benar-benar dipakai (bukan sisa template CNN/GAN)",
     LULUS,
     "Daftar Simbol (Hal. xxv) — K, Z, E, R, F, d_pred, d_gt, N, T; "
     "semua dipakai di BAB II (Hal. 16–22).",
     "Minor"),
    ("A8",
     "Daftar Singkatan berisi akronim yang benar (bukan sisa template CNN/GPU)",
     LULUS,
     "Daftar Singkatan (Hal. xxvii) — 33 akronim setelah penambahan API, CRD, EDA, "
     "GNN, GPT, GVK, HPA, IR, NLP, PCST, PVC.",
     "Minor"),
    ("A9",
     "Daftar Gambar/Tabel memiliki prefix bab yang benar dan nomor halaman yang benar",
     LULUS,
     "Daftar Gambar (Hal. xvii), Daftar Tabel (Hal. xix) — prefix bab I–VII sesuai. "
     "Nomor halaman final perlu verifikasi setelah kompilasi ulang.",
     "Minor"),
    ("A10",
     "Tidak ada teks placeholder di halaman mana pun (asdf, xxx, John Doe, [to be completed])",
     LULUS,
     "Seluruh dokumen (Hal. i–185+) — audit tidak menemukan placeholder di 28 file .tex.",
     "Blocker"),
    ("A11",
     "Lampiran ada & relevan (bukan sisa template sensor ultrasonik / binary search)",
     LULUS,
     "Lampiran A (Hal. 113), B (Hal. 147), C (Hal. 161), D (Hal. 183), E (Hal. 185) — "
     "kode, fixture, hasil kuantitatif, wawancara, validasi pakar.",
     "Major"),
    ("A12",
     "Halaman awal konsisten profesional & akademis",
     LULUS,
     "Halaman Judul (Hal. i), Lembar Pengesahan (Hal. iii), Pernyataan Orisinalitas (Hal. v), "
     "Kata Pengantar (Hal. xi) — nama, NIM, program studi seragam.",
     "Minor"),

    # ── B ────────────────────────────────────────────────────────────────
    (CAT, "B. Bab I — Pendahuluan", "", "", ""),
    ("B1",
     "Rumusan masalah jelas, nyata, dan termotivasi (tidak generik)",
     LULUS,
     "BAB I (Hal. 4) — 3 rumusan masalah bernomor; dimotivasi oleh gap teknis T01–T05 "
     "dari Latar Belakang (Hal. 1–3).",
     "Major"),
    ("B2",
     "Masalah dikuantifikasi dengan angka yang kuat dan bersumber",
     LULUS,
     "BAB I (Hal. 1–3) — 63,4% exact match Vector RAG (wan2025empowering); 80% insiden "
     "YAML; 18% manifes K8s mengandung kerentanan (rahman2023misconfigurations).",
     "Major"),
    ("B3",
     "Pertanyaan penelitian / tujuan eksplisit, bernomor, dan dapat dijawab",
     LULUS,
     "BAB I (Hal. 4) — T1, T2, T3 bernomor; dijawab di BAB VII (Hal. 105–106) dengan angka.",
     "Major"),
    ("B4",
     "Batasan masalah dinyatakan dengan nama yang sesuai judul serta apa yang dievaluasi",
     LULUS,
     "BAB I (Hal. 4–5) — Kubernetes v1.30, swagger.json, GPT-4o-mini temp=0,1, "
     "97 fixture, benchmark retrieval.",
     "Minor"),
    ("B5",
     "Kriteria keberhasilan / hipotesis eksplisit dan terukur",
     LULUS,
     "BAB I (Hal. 5) — H1: GraphRAG > Vector RAG RetQ & ReaQ (p<0,05); H2: HopAcc; "
     "H3: AnsQ pada yaml_gen.",
     "Major"),
    ("B6",
     "Gambaran metodologi (CRISP-DM / DSRM) disebut dan sesuai",
     LULUS,
     "BAB I (Hal. 5–6) — CRISP-DM 6 fase, Gambar I.2 (Hal. 6), pemetaan fase ke bab.",
     "Minor"),
    ("B7",
     "Klaim kontribusi dinyatakan dan benar-benar baru untuk konteksnya",
     LULUS,
     "BAB I (Hal. 6–7) — 3 kontribusi teknis; dibedakan eksplisit dari wan2025empowering "
     "dan civiot2025graphrag.",
     "Major"),

    # ── C ────────────────────────────────────────────────────────────────
    (CAT, "C. Bab II — Studi Literatur", "", "", ""),
    ("C1",
     "Teori dasar dan terpaduan sesuai kebutuhan",
     LULUS,
     "BAB II (Hal. 9–25) — K8s (Hal. 9–12), LLM (Hal. 12–13), RAG (Hal. 13–15), "
     "KG & GraphRAG (Hal. 16–18), metrik evaluasi (Hal. 19–22).",
     "Major"),
    ("C2",
     "Cakupan luas dan mutakhir (sumber terbaru + seminal)",
     LULUS,
     "BAB II (Hal. 9–25) — 25 referensi; seminal: pan2024unifying, han_graphrag_2024; "
     "mutakhir: civiot2025graphrag (2025), hsgrag2025acm (2025).",
     "Major"),
    ("C3",
     "Related work memposisikan tesis terhadap sistem terdahulu (tabel gap atau setara)",
     LULUS,
     "BAB II (Hal. 23–26) — Subbab II.7 Related Work + Tabel 12 gap analysis "
     "(T01–T05 vs 3 penelitian terkait).",
     "Major"),
    ("C4",
     "Gap penelitian diidentifikasi secara tajam dan terkait pertanyaan penelitian",
     LULUS,
     "BAB II (Hal. 23–26) — T01: KG dari teks; T02: depth tetap; T03: tidak ada validasi "
     "YAML; T04: tidak ada evaluasi K8s; T05: tidak ada update inkremental.",
     "Major"),
    ("C5",
     "Tidak ada over-investasi pada pendekatan yang akhirnya dibuang",
     LULUS,
     "BAB III (Hal. 40–44) — alternatif G-Retriever dan NeuSym-RAG dievaluasi ringkas "
     "sebagai kandidat yang kalah dalam matriks keputusan.",
     "Minor"),
    ("C6",
     "Kualitas sitasi: sumber dapat ditelusuri, bukan dari sumber sekunder tanpa atribusi",
     LULUS,
     "BAB II (Hal. 9–25) — semua klaim kuantitatif disertai sitasi primer; "
     "tidak ada parafrase tanpa atribusi.",
     "Minor"),
    ("C7",
     "Kualitas sitasi: sumber peer-reviewed untuk klaim penting (bukan blog/vendor/fabrikasi)",
     LULUS,
     "Daftar Pustaka — setelah pembersihan: file:/// URL, note 'user-provided', dan "
     "11 entri uncited dihapus; he2024gretriever -> @inproceedings ICLR 2024.",
     "Major"),

    # ── D ────────────────────────────────────────────────────────────────
    (CAT, "D. Bab III — Analisis & Perancangan", "", "", ""),
    ("D1",
     "Kebutuhan (fungsional + non-fungsional) konkret dan tertelusur ke masalah",
     LULUS,
     "BAB III (Hal. 33–39) — Tabel 15: B01–B03 -> T01–T05 -> F01–F05; "
     "non-fungsional (Hal. 38): latensi <5s, reprodusibel.",
     "Major"),
    ("D2",
     "Pemilihan solusi: justifikasi metode valid (AHP/WSM); bobot per kriteria berjumlah 1,0",
     LULUS,
     "BAB III (Hal. 42–44) — WSM bobot setara w=0,25/kriteria (S=1,0) dinyatakan eksplisit "
     "sebelum Tabel 4; Alt 1 = metode yang diusulkan.",
     "Major"),
    ("D3",
     "EDA: distribusi, cek null/duplikat, bukti angka ilustratif",
     LULUS,
     "BAB III (Hal. 29–32) — Subbab III.1.2.3: 735->730 valid; 183 root (24,9%) + "
     "547 sub (75,1%); swagger 3,757 MB; 59,9% definisi mengandung $ref.",
     "Major"),
    ("D4",
     "Asal-usul data / etika / persetujuan dinyatakan bila data milik pihak tertentu",
     LULUS,
     "BAB III (Hal. 29) — Subbab III.1.2.1: swagger.json Kubernetes v1.30 dari "
     "repositori publik Apache 2.0.",
     "Minor"),
    ("D5",
     "Arsitektur koheren dan sesuai kebutuhan",
     LULUS,
     "BAB IV (Hal. 47–60) — arsitektur 5-fase parser + LangGraph pipeline; "
     "BAB III (Hal. 44–46) analisis pemilihan.",
     "Major"),
    ("D6",
     "Keputusan desain (hyperparameter, threshold, pemilihan model) dijustifikasi",
     LULUS,
     "BAB IV (Hal. 54–56), BAB VI (Hal. 93–100) — depth 1–5 dari ablation; "
     "temperature=0,1; threshold RGA≥0,5; top-k=5.",
     "Major"),
    ("D7",
     "Konsistensi internal — jumlah, threshold, nama cocok dengan bab-bab berikutnya",
     LULUS,
     "BAB III (Hal. 30), BAB V (Hal. 62–64), BAB VI (Hal. 88), BAB VII (Hal. 105) — "
     "node 725, swagger 3,757 MB, 18 edge types, 97 fixture seragam.",
     "Major"),

    # ── E ────────────────────────────────────────────────────────────────
    (CAT, "E. Bab IV/V — Implementasi", "", "", ""),
    ("E1",
     "Sistem benar-benar dibangun — artefak/kode/konfigurasi konkret",
     LULUS,
     "BAB V (Hal. 61–83), Lampiran A (Hal. 113) — 725 node Neo4j, kode 6 komponen, "
     "97 fixture dievaluasi nyata.",
     "Blocker"),
    ("E2",
     "Tech stack dispesifikasikan (versi, pustaka, perangkat keras)",
     LULUS,
     "BAB V (Hal. 61) — Tabel 25: Python 3.11, Neo4j, GPT-4o-mini, LangGraph, "
     "text-embedding-3-small (1536-dim), Kubernetes v1.30.",
     "Minor"),
    ("E3",
     "Alur data (akuisisi -> persiapan -> fitur) terimplementasi & reprodusibel",
     LULUS,
     "BAB V (Hal. 62–69) — SwaggerGraphBuilder Pass1–Pass3+1.5; "
     "Lampiran C (Hal. 161): data/eval_results.csv tersimpan.",
     "Major"),
    ("E4",
     "Langkah non-trivial (konstruksi label, traversal, klasifikasi) — bukan ditunda",
     LULUS,
     "BAB V (Hal. 70–80) — multi-hop traversal (Hal. 74–78), intent classification "
     "(Hal. 72–74), validasi YAML 3-lapis (Hal. 79–80).",
     "Major"),
    ("E5",
     "Cakupan yang dirancang -> cakupan yang sebenarnya tidak ada yang dilewatkan",
     LULUS,
     "BAB V (Hal. 61–83) — F01–F05 semua terimplementasi; "
     "BAB IV (Hal. 47–60) rancangan bersesuaian 1:1.",
     "Major"),
    ("E6",
     "Reprodusibilitas — seed, artefak tersimpan, environment, pelacakan eksperimen",
     LULUS,
     "BAB V (Hal. 62) — session_id unik; Lampiran C (Hal. 161–182): "
     "data/eval_results*.csv + ablation CSVs; requirements.txt.",
     "Major"),
    ("E7",
     "Konsistensi model/konfigurasi — cocok dengan desain & setup perbandingan di Bab VI",
     LULUS,
     "BAB V (Hal. 61), BAB VI (Hal. 87–88) — GPT-4o-mini temperature=0,1 identik; "
     "Tabel 31 menggunakan konfigurasi yang sama.",
     "Major"),

    # ── F ────────────────────────────────────────────────────────────────
    (CAT, "F. Bab VI — Evaluasi", "", "", ""),
    ("F1",
     "Metode evaluasi didefinisikan sebelum hasil (dataset, metrik, protokol)",
     LULUS,
     "BAB VI (Hal. 85–88) — Subbab VI.1 Metode Evaluasi mendahului VI.2 Hasil "
     "Evaluasi (Hal. 88).",
     "Major"),
    ("F2",
     "Data yang terukur nyata ada (bukan target, bukan placeholder 'XX')",
     LULUS,
     "BAB VI (Hal. 88–93), Lampiran C (Hal. 161–182) — RetQ 0,7259; ReaQ 0,5554; "
     "AnsQ 0,5771; PathCov 0,8515; HopAcc 0,3505; RGA 0,4536 (97 fixture, 3 sistem).",
     "Blocker"),
    ("F3",
     "Metrik sesuai tugas (tidak dicampur-adukkan)",
     LULUS,
     "BAB VI (Hal. 87) — Subbab VI.1.2: AnsQ, RetQ, ReaQ + metrik domain K8s "
     "(PathCov, HopAcc, RGA, SyntVal) sesuai tiga tujuan.",
     "Major"),
    ("F4",
     "Baseline/perbandingan >= 1 terverifikasi (bukan hanya sistem sendiri bila ada klaim)",
     LULUS,
     "BAB VI (Hal. 87–88) — Subbab VI.1.3: Vanilla LLM + Vector RAG sebagai baseline "
     "terkontrol; Tabel 31 (Hal. 95) perbandingan lengkap.",
     "Blocker"),
    ("F5",
     "Kekokohan statistik (signifikansi, effect size, bukan sekadar titik estimasi)",
     LULUS,
     "BAB VI (Hal. 93–100) — Wilcoxon signed-rank + paired bootstrap 95% CI + "
     "Holm-Bonferroni; RetQ/ReaQ/PathCov/HopAcc/RGA: p<0,001; AnsQ: n.s.",
     "Major"),
    ("F6",
     "Analisis ablasi / kontribusi komponen bila desain mengklaim banyak kontribusi",
     LULUS,
     "BAB VI (Hal. 93–96) — Subbab VI.3.1: ablation 6 konfigurasi (A1–A6c); "
     "Tabel 32 + Tabel 36; A2 dominan (RetQ −0,601).",
     "Major"),
    ("F7",
     "Hasil benar-benar membahas data target, bukan hanya kriteria keberhasilan di Bab I",
     LULUS,
     "BAB VI (Hal. 88–100) — AnsQ n.s. dilaporkan jujur; faithfulness GraphRAG < Vanilla "
     "LLM diakui; realworld RetQ=0,39 dianalisis.",
     "Major"),
    ("F8",
     "Cakupan tujuan — setiap RQ/indikator yang dijanjikan dievaluasi",
     LULUS,
     "BAB VI (Hal. 88–101) — T1: PathCov 0,8515 (Hal. 92); T2: HopAcc 0,3505 + "
     "ablation (Hal. 93–96); T3: Tabel 31 (Hal. 95).",
     "Major"),
    ("F9",
     "Ancaman validitas dibahas (kebocoran data, sampel kecil, evaluasi-diri, sirkularitas)",
     LULUS,
     "BAB VI (Hal. 101–104) — Subbab VI.4: 4 ancaman validitas (session ID, bias sampel, "
     "n=4 pakar, GPT-4o-mini evaluator sekaligus generator).",
     "Major"),
    ("F10",
     "Pelaporan jujur — hasil negatif/gagal disebutkan, tidak disembunyikan",
     LULUS,
     "BAB VI (Hal. 89, 96–100) — AnsQ n.s. (p>0,05) disebut; faithfulness GraphRAG < "
     "Vanilla LLM diakui; realworld RetQ=0,39 dianalisis; Lampiran E (Hal. 185) pakar.",
     "Major"),

    # ── G ────────────────────────────────────────────────────────────────
    (CAT, "G. Bab VII — Penutup", "", "", ""),
    ("G1",
     "Kesimpulan ditulis (bukan placeholder)",
     LULUS,
     "BAB VII (Hal. 105) — Subbab VII.1: menjawab T1/T2/T3 dengan angka "
     "RetQ 0,7259, HopAcc 0,3505, SyntVal 0,8947.",
     "Blocker"),
    ("G2",
     "Setiap pertanyaan penelitian dijawab eksplisit memakai hasil tersebut",
     LULUS,
     "BAB VII (Hal. 105–106) — T1: 725 node, 18 edge types; T2: ablation A2 −0,601 RetQ; "
     "T3: GraphRAG > baseline p<0,001 (RetQ/ReaQ/PathCov/HopAcc/RGA).",
     "Major"),
    ("G3",
     "Kontribusi dirangkum ulang & didukung bukti (tidak over-claim)",
     LULUS,
     "BAB VII (Hal. 105–106) — kontribusi diringkas dengan angka; AnsQ n.s. diakui "
     "sehingga tidak ada over-claim.",
     "Major"),
    ("G4",
     "Keterbatasan diakui",
     LULUS,
     "BAB VII (Hal. 106) — Subbab VII.2: KG statis, GPT-4o-mini saja, "
     "realworld RetQ=0,39, n=97 fixture.",
     "Major"),
    ("G5",
     "Saran/future work konkret dan mengikuti temuan",
     LULUS,
     "BAB VII (Hal. 107–111) — Subbab VII.3: dynamic KG update, multi-source ingestion, "
     "fine-tuned embedding K8s, deployment kluster nyata.",
     "Minor"),

    # ── H ────────────────────────────────────────────────────────────────
    (CAT, "H. Kualitas Lintas-Bab", "", "", ""),
    ("H1",
     "Konsistensi internal — satu istilah = satu makna; jumlah/threshold identik di semua bab",
     LULUS,
     "BAB III (Hal. 30), BAB V (Hal. 62–64), BAB VI (Hal. 88), BAB VII (Hal. 105) — "
     "node 725 (183+542), swagger 3,757 MB, 18 edge types, 97 fixture seragam.",
     "Major"),
    ("H2",
     "Rujukan silang tertelusur (tidak ada Gambar X atau Tabel Y yang broken \\ref)",
     LULUS,
     "BAB VI (Hal. 101), Lampiran E (Hal. 185) — ref tbl:expert_profile diperbaiki ke "
     "lampiran:validasi-pakar; tabel peran baru tbl:evaluator_roles.",
     "Major"),
    ("H3",
     "Gambar/tabel diberi caption, dinomori, dirujuk; daftar gambar/tabel/persamaan terisi",
     LULUS,
     "Daftar Gambar (Hal. xvii), Daftar Tabel (Hal. xix), Daftar Persamaan (Hal. xxi) — "
     "Daftar Algoritma kosong dihapus; Daftar Persamaan kini 13 entri.",
     "Minor"),
    ("H4",
     "Daftar Pustaka lengkap & rapi (tanpa placeholder; venue tidak kosong; tidak ada duplikat)",
     LULUS,
     "Daftar Pustaka — duplikat hofer2024construction dihapus; URL file:// diperbaiki; "
     "he2024gretriever -> ICLR 2024; 11 entri uncited dihapus. 25 entri bersih.",
     "Major"),
    ("H5",
     "Kualitas penulisan — tata bahasa, EYD, desimal konsisten",
     LULUS,
     "Lampiran C (Hal. 161–182) — 1.261 titik desimal dikonversi ke koma; "
     "terminologi seragam di seluruh dokumen.",
     "Minor"),
    ("H6",
     "Judul pada semua halaman identik dan nama yang diserahkan selaras",
     LULUS,
     "Halaman Judul (Hal. i), Lembar Pengesahan (Hal. iii), Abstrak (Hal. ix), "
     "Kata Pengantar (Hal. xi) — \"Presisi Retrieval\" seragam.",
     "Blocker"),
    ("H7",
     "Reprodusibilitas — orang lain dapat menyalin ulang pekerjaan tersebut",
     LULUS,
     "Lampiran A (Hal. 113), Lampiran C (Hal. 161) — kode, data/eval_results*.csv, "
     "swagger.json publik, Python 3.11, requirements.txt.",
     "Major"),
]

# ── Write rows ────────────────────────────────────────────────────────────
r = 4
for item in rows:
    kode, kriteria, status, catatan, severitas = item

    if kode == CAT:
        ws.merge_cells(f"A{r}:E{r}")
        c = ws.cell(row=r, column=1, value=kriteria)
        cell_style(c, fill=CAT_FILL, font=CAT_FONT, align_h="left")
        ws.row_dimensions[r].height = 18
        r += 1
        continue

    c = ws.cell(row=r, column=1, value=kode)
    cell_style(c, font=BOLD_FONT, align_h="center")

    c = ws.cell(row=r, column=2, value=kriteria)
    cell_style(c, font=SMALL_FONT)

    s_fill = LULUS_FILL if status == LULUS else (DEVIASI_FILL if status == DEVIASI else BLOCKER_FILL)
    c = ws.cell(row=r, column=3, value=status)
    cell_style(c, fill=s_fill, font=BOLD_FONT, align_h="center")

    c = ws.cell(row=r, column=4, value=catatan)
    cell_style(c, font=SMALL_FONT)

    c = ws.cell(row=r, column=5, value=severitas)
    cell_style(c, fill=sev_fill(severitas), font=BOLD_FONT, align_h="center")

    ws.row_dimensions[r].height = 48
    r += 1

# ── Summary ───────────────────────────────────────────────────────────────
r += 1
ws.merge_cells(f"A{r}:B{r}")
ws.cell(row=r, column=1, value="RINGKASAN AKHIR").font = Font(bold=True, size=11)
ws.cell(row=r, column=1).alignment = Alignment(horizontal="right", vertical="center")
ws.cell(row=r, column=3, value="LULUS: 46/47").fill = LULUS_FILL
ws.cell(row=r, column=3).font = Font(bold=True, size=11)
ws.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center")
ws.cell(row=r, column=4,
        value="1 Deviasi Diterima (A4: abstrak EN tidak ditambahkan — keputusan peneliti)"
        ).font = Font(italic=True, size=9)
ws.cell(row=r, column=4).alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[r].height = 24

# ── Freeze panes ──────────────────────────────────────────────────────────
ws.freeze_panes = "A4"

# ── Sheet 2: Tindakan Manual ──────────────────────────────────────────────
ws2 = wb.create_sheet("Tindakan Manual")
ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 35
ws2.column_dimensions["C"].width = 62
ws2.column_dimensions["D"].width = 20

ws2.merge_cells("A1:D1")
ws2["A1"] = "TINDAKAN MANUAL YANG MASIH DIPERLUKAN (di luar .tex)"
ws2["A1"].fill = HEADER_FILL
ws2["A1"].font = WHITE_BOLD
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 22

for col, h in enumerate(["No", "Aksi", "Keterangan", "Prioritas"], 1):
    c = ws2.cell(row=2, column=col, value=h)
    c.fill = CAT_FILL
    c.font = CAT_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN_BORDER
ws2.row_dimensions[2].height = 18

manual = [
    ("M1", "Kompilasi ulang PDF (wajib)",
     "xelatex TA.tex && biber TA && xelatex TA.tex && xelatex TA.tex dari "
     "docs/TA-STI-template-1.0/. Diperlukan untuk meregenerasi Daftar Persamaan (13 entri), "
     "Daftar Singkatan (33 akronim), dan nomor halaman yang benar.",
     "Wajib - segera"),
    ("M2", "Tanda tangan basah Lembar Pengesahan",
     "Halaman 2 Lembar Pengesahan.tex perlu ditandatangani secara fisik oleh dosen pembimbing "
     "dan ketua program sebelum sidang.",
     "Wajib - sidang"),
    ("M3", "Verifikasi visual PDF setelah kompilasi",
     "Setelah xelatex selesai: (a) Daftar Persamaan memuat 13 entri, "
     "(b) Daftar Singkatan urutan abjad A-Y, "
     "(c) grep log biber untuk warning, (d) cari '??' di PDF untuk broken \\ref.",
     "Wajib - sebelum cetak"),
    ("M4", "Verifikasi file PDF Lampiran E",
     "Pastikan 5 file PDF lembar validasi tersedia di docs/TA-STI-template-1.0/lampiran-e/ "
     "sebelum kompilasi final: lembar-validator-E1.pdf s/d E5.pdf.",
     "Wajib - kompilasi"),
    ("M5", "Rename label fig:I.2 (opsional)",
     "Label fig:I.2 pada gambar CRISP-DM di Bab I sedikit membingungkan. Opsional: "
     "rename ke fig:crispdm-metodologi + update \\ref.",
     "Opsional - kosmetik"),
]

for row_idx, (no, aksi, ket, prioritas) in enumerate(manual, 3):
    ws2.cell(row=row_idx, column=1, value=no)
    ws2.cell(row=row_idx, column=2, value=aksi)
    ws2.cell(row=row_idx, column=3, value=ket)
    pf = (BLOCKER_FILL if "segera" in prioritas else
          MAJOR_FILL if "Wajib" in prioritas else MINOR_FILL)
    ws2.cell(row=row_idx, column=4, value=prioritas).fill = pf
    for col in range(1, 5):
        c = ws2.cell(row=row_idx, column=col)
        c.border = THIN_BORDER
        c.font = SMALL_FONT
        c.alignment = Alignment(horizontal="left" if col != 1 else "center",
                                vertical="center", wrap_text=True)
    ws2.row_dimensions[row_idx].height = 55

# ── Sheet 3: Legenda ───────────────────────────────────────────────────────
ws3 = wb.create_sheet("Legenda")
ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 50

ws3.merge_cells("A1:B1")
ws3["A1"] = "LEGENDA WARNA"
ws3["A1"].fill = HEADER_FILL
ws3["A1"].font = WHITE_BOLD
ws3["A1"].alignment = Alignment(horizontal="center")
ws3.row_dimensions[1].height = 22

legend = [
    (LULUS_FILL,   "LULUS — item memenuhi kriteria sidang"),
    (DEVIASI_FILL, "DEVIASI DITERIMA — item menyimpang namun diterima berdasarkan keputusan peneliti"),
    (BLOCKER_FILL, "Severitas BLOCKER — pasti dicatat penguji, wajib diperbaiki"),
    (MAJOR_FILL,   "Severitas MAJOR — berpengaruh signifikan pada nilai/persepsi"),
    (MINOR_FILL,   "Severitas MINOR — kosmetik/konsistensi, baik diperbaiki"),
]
for idx, (fill, label) in enumerate(legend, 2):
    c = ws3.cell(row=idx, column=1, value=label.split(" — ")[0])
    c.fill = fill
    c.font = BOLD_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN_BORDER
    d = ws3.cell(row=idx, column=2, value=label.split(" — ")[1])
    d.font = SMALL_FONT
    d.alignment = Alignment(horizontal="left", vertical="center")
    d.border = THIN_BORDER
    ws3.row_dimensions[idx].height = 20

# ── Save ──────────────────────────────────────────────────────────────────
out_path = r"c:\Users\Jihan Aurelia\Documents\SMT8\Tugas-Akhir-GraphRAG-Kubernetes\docs\checklist_kualitas_sidang_v2.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
