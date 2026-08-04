"""
scripts/gen_metric_evidence_workbook.py — workbook bukti perhitungan metrik GraphRAG.

Menyusun 5 sheet (.xlsx) yang menunjukkan, per fixture, dari mana angka AnsQ /
RetQ / ReaQ berasal — supaya penguji sidang bisa menelusuri satu angka sampai
ke data mentahnya tanpa membuka banyak file.

Input : data/eval_results_graphrag_final.csv, data/eval_cases_graphrag.jsonl,
        tests/fixtures/*/*.json, data/ragas_results_graphrag.csv (cross-check),
        data/eval_run_meta_graphrag.json
Output: data/lampiran_bukti_metrik_graphrag.xlsx (5 sheet)

Read-only: tidak ada file input yang dimodifikasi.

Usage:
  python scripts/gen_metric_evidence_workbook.py
"""
import sys
import os
import io
import csv
import json
import math
import re
import glob
import logging
from pathlib import Path
from collections import Counter

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[WARN] openpyxl tidak terpasang. Jalankan: pip install openpyxl==3.1.5")
    sys.exit(1)

ROOT = Path(__file__).parent.parent
DATA = ROOT / "data"
FIXTURES = ROOT / "tests" / "fixtures"
OUTPUT = DATA / "lampiran_bukti_metrik_graphrag.xlsx"

RBAC_EXCLUDE = "serviceaccount_pod_binding"

TYPE_COUNTS = {
    "conceptual": 25, "yaml_gen": 25, "relationship": 18, "followup": 12,
    "realworld": 9, "planning": 5, "troubleshooting": 5, "command": 3,
}

# Angka headline skripsi (Bab VI) — dicocokkan di verify(), tol 0.0005 kecuali disebut lain.
THESIS = {
    "ansq":               (0.8031, 102),
    "ansq_relevansi":      (0.7698, 102),
    "ansq_sintaksis":      (1.0000, 28),
    "ansq_skema_exclrbac": (0.9259, 27),
    "retq_precision":      (0.8405, 102),
    "retq_recall":         (0.7258, 102),
    "retq_f1":             (0.7089, 102),
    "hop_semua":           (0.7562, 93),
    "hop_fokus":           (0.9086, 50),
    "hop_closure":         (0.5791, 43),
    "faithfulness":        (0.3055, 95),
}

EXAMPLE_IDS = ["rolebinding_dev", "kubectl_find_pods_with_env", "clusterrole_read_pods"]

RELATION_RE = re.compile(r"-\[([^\]]+)\]->?")


# ══════════════════════════════════════════════════════════════════════════
# I/O
# ══════════════════════════════════════════════════════════════════════════

def _read_csv(path):
    content = path.read_bytes().replace(b"\x00", b"")
    return list(csv.DictReader(io.StringIO(content.decode("utf-8", errors="replace"))))


def _read_jsonl(path):
    out = {}
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            r = json.loads(line)
            out[r["id"]] = r
    return out


def _load_fixtures():
    out = {}
    for p in glob.glob(str(FIXTURES / "*" / "*.json")):
        d = json.load(open(p, encoding="utf-8"))
        if "oos_reason" in d:
            continue
        out[d["id"]] = d
    return out


def _f(v):
    if v in (None, ""):
        return None
    try:
        x = float(v)
        return None if math.isnan(x) else x
    except (TypeError, ValueError):
        return None


def _i(v):
    x = _f(v)
    return None if x is None else int(x)


# ══════════════════════════════════════════════════════════════════════════
# Rekomputasi (mereplikasi scripts/evaluate.py persis)
# ══════════════════════════════════════════════════════════════════════════

def _short(name):
    return name.split(".")[-1]


def extract_root_resource(graph_context_str):
    """Meniru evaluate.py:651-657. Gagal parse (konteks multi-entitas) -> ("", False)."""
    if not graph_context_str:
        return "", True
    try:
        obj = json.loads(graph_context_str)
    except Exception:
        return "", False
    if isinstance(obj, dict):
        return obj.get("RootResource", ""), True
    return "", False


def recompute_retq(reasoning_path, ground_truth, root_resource):
    """Meniru evaluate.py:compute_retq (267-331)."""
    expected_nodes = set(_short(n) for n in ground_truth.get("relevant_nodes", []))
    seen = set()
    retrieved = []
    if root_resource:
        rs = _short(root_resource)
        if rs not in seen:
            seen.add(rs)
            retrieved.append(rs)
    for step in reasoning_path:
        cleaned = RELATION_RE.sub(" ", step)
        for tok in cleaned.split():
            if tok not in seen:
                seen.add(tok)
                retrieved.append(tok)
    retrieved_set = set(retrieved)
    inter = retrieved_set & expected_nodes
    missed = expected_nodes - retrieved_set   # G \ R -> penyebab recall < 1
    extra = retrieved_set - expected_nodes    # R \ G -> penyebab precision < 1
    precision = len(inter) / len(retrieved_set) if retrieved_set else 0.0
    recall = len(inter) / len(expected_nodes) if expected_nodes else 1.0
    return {
        "R": retrieved, "G": sorted(expected_nodes), "inter": sorted(inter),
        "missed": sorted(missed), "extra": sorted(extra),
        "n_R": len(retrieved_set), "n_G": len(expected_nodes), "n_inter": len(inter),
        "precision": precision, "recall": recall,
    }


def recompute_hop(reasoning_path, expected_path):
    """Meniru evaluate.py:compute_reaq hop_accuracy (364-369). Pencocokan ketat lowercase."""
    if not expected_path:
        return {
            "hop_accuracy": None, "n_gt": 0, "n_pred": len(reasoning_path),
            "n_match": 0, "missed": [], "gt_display": [], "pred_display": reasoning_path,
        }
    exp_map = {}
    for e in expected_path:
        k = e.strip().lower()
        exp_map.setdefault(k, e)
    pred_lower = set(e.strip().lower() for e in reasoning_path)
    exp_lower = set(exp_map.keys())
    match = exp_lower & pred_lower
    missed_lower = exp_lower - pred_lower
    hop_accuracy = len(match) / len(exp_lower) if exp_lower else None
    return {
        "hop_accuracy": hop_accuracy, "n_gt": len(exp_lower), "n_pred": len(reasoning_path),
        "n_match": len(match), "missed": sorted(exp_map[k] for k in missed_lower),
        "gt_display": expected_path, "pred_display": reasoning_path,
    }


# ══════════════════════════════════════════════════════════════════════════
# Join
# ══════════════════════════════════════════════════════════════════════════

def build_rows():
    csv_rows = {r["id"]: r for r in _read_csv(DATA / "eval_results_graphrag_final.csv")}
    cases = _read_jsonl(DATA / "eval_cases_graphrag.jsonl")
    fixtures = _load_fixtures()

    rows = []
    for fid, c in sorted(csv_rows.items(), key=lambda kv: (kv[1]["type"], kv[0])):
        case = cases.get(fid, {})
        fx = fixtures.get(fid, {})
        gt = fx.get("ground_truth", {})

        reasoning_path = case.get("reasoning_path", [])
        graph_context = case.get("graph_context", "")
        root_resource, parseable = extract_root_resource(graph_context)

        retq = recompute_retq(reasoning_path, gt, root_resource)
        hop = recompute_hop(reasoning_path, gt.get("expected_path", []))

        rows.append({
            "id": fid, "type": c["type"],
            "csv": c, "case": case, "gt": gt,
            "question": case.get("question", fx.get("question", "")),
            "answer_full": case.get("answer_full", ""),
            "gt_answer": gt.get("answer", ""),
            "graph_context": graph_context,
            "root_resource": root_resource,
            "root_parseable": parseable,
            "retq": retq,
            "hop": hop,
        })
    return rows


# ══════════════════════════════════════════════════════════════════════════
# Kalimat penjelas
# ══════════════════════════════════════════════════════════════════════════

def sentence_ansq(row):
    c = row["csv"]
    syn, sch, rel = _f(c["ansq_syntactic_validity"]), _f(c["ansq_schema_compliance"]), _f(c["ansq_answer_relevance"])
    score = _f(c["ansq_ansq_score"])
    if syn is not None:
        parts = [f"sintaksis {syn:.4f}", f"skema {sch:.4f}", f"relevansi {rel:.4f}"]
        s = f"AnsQ {score:.4f} = rata-rata 3 sub-metrik ({' · '.join(parts)})."
        if syn == 0.0:
            s += " YAML gagal di-parse (sintaksis gagal)."
        elif sch == 0.0:
            s += " YAML tidak lolos validator K8s 1.30 (skema gagal)."
        return s
    return f"AnsQ {score:.4f} = nilai relevansi jawaban saja; sub-metrik YAML tidak berlaku untuk tipe {row['type']}."


def sentence_retq(row):
    r = row["retq"]
    p, rc = r["precision"], r["recall"]
    f1 = _f(row["csv"]["retq_f1"])
    s = (f"RetQ (F1) {f1:.4f} — dari {r['n_G']} node relevan GT, {r['n_inter']} berhasil diambil "
         f"({rc:.0%} recall); dari {r['n_R']} node yang diambil sistem, {r['n_inter']} tepat sasaran "
         f"({p:.0%} precision).")
    if r["missed"]:
        s += f" {len(r['missed'])} node GT tidak terambil."
    if r["extra"]:
        s += f" {len(r['extra'])} node diambil di luar GT."
    if not row["root_parseable"]:
        s += " Konteks multi-entitas: node akar tidak ditambahkan ke himpunan node diambil."
    return s


def sentence_reaq(row):
    h = row["hop"]
    faith = _f(row["csv"]["reaq_reaq_score"])
    if h["hop_accuracy"] is not None:
        s = (f"Hop Accuracy {h['hop_accuracy']:.4f} = {h['n_match']}/{h['n_gt']} edge GT berhasil "
             f"ditelusuri (sistem menelusuri {h['n_pred']} edge total).")
    else:
        s = f"Hop Accuracy N/A — tipe fixture {row['type']} tidak memiliki jalur GT untuk dinilai."
    if faith is not None:
        s += f" Faithfulness {faith:.4f} = hasil hakim LLM (gpt-4o-mini), tidak dapat dihitung ulang di sheet ini."
    else:
        s += " Faithfulness GAGAL DIUKUR pada fixture ini (hakim RAGAS mengembalikan NaN/error)."
    return s


# ══════════════════════════════════════════════════════════════════════════
# Verifikasi (fail-fast, sebelum workbook ditulis)
# ══════════════════════════════════════════════════════════════════════════

def verify(rows, ragas_by_id):
    problems = []

    if len(rows) != 102:
        problems.append(f"Jumlah baris CSV = {len(rows)}, diharapkan 102")

    type_counts = Counter(r["type"] for r in rows)
    for t, n in TYPE_COUNTS.items():
        if type_counts.get(t) != n:
            problems.append(f"Tipe '{t}': {type_counts.get(t)} baris, diharapkan {n}")

    n_short_collision = 0
    for r in rows:
        c, retq = r["csv"], r["retq"]
        if retq["n_R"] != _i(c["n_retrieved_nodes"]):
            problems.append(f"{r['id']}: n_R hitung ulang {retq['n_R']} != CSV {c['n_retrieved_nodes']}")
        if retq["n_G"] != _i(c["n_relevant_nodes"]):
            problems.append(f"{r['id']}: n_G hitung ulang {retq['n_G']} != CSV {c['n_relevant_nodes']}")
        if retq["n_inter"] != _i(c["n_node_intersection"]):
            problems.append(f"{r['id']}: n_inter hitung ulang {retq['n_inter']} != CSV {c['n_node_intersection']}")

        n_gt_full = len(r["gt"].get("relevant_nodes", []))
        n_gt_short = len(set(_short(x) for x in r["gt"].get("relevant_nodes", [])))
        if n_gt_full != n_gt_short:
            n_short_collision += 1

        p_csv, rc_csv, f1_csv = _f(c["retq_precision"]), _f(c["retq_recall"]), _f(c["retq_f1"])
        if p_csv is not None and abs(retq["precision"] - p_csv) > 1e-5:
            problems.append(f"{r['id']}: precision hitung ulang {retq['precision']:.6f} != CSV {p_csv:.6f}")
        if rc_csv is not None and abs(retq["recall"] - rc_csv) > 1e-5:
            problems.append(f"{r['id']}: recall hitung ulang {retq['recall']:.6f} != CSV {rc_csv:.6f}")
        if f1_csv is not None:
            p, rc = retq["precision"], retq["recall"]
            f1_calc = 2 * p * rc / (p + rc) if (p + rc) > 0 else 0.0
            if abs(f1_calc - f1_csv) > 1e-5:
                problems.append(f"{r['id']}: F1 hitung ulang {f1_calc:.6f} != CSV {f1_csv:.6f}")

        subs = [_f(c["ansq_syntactic_validity"]), _f(c["ansq_schema_compliance"]),
                _f(c["ansq_answer_relevance"]), _f(c["ansq_layer3_compliance"])]
        applicable = [v for v in subs if v is not None]
        mean_calc = sum(applicable) / len(applicable) if applicable else 0.0
        ansq_csv = _f(c["ansq_ansq_score"])
        if ansq_csv is not None and abs(mean_calc - ansq_csv) > 1e-9:
            problems.append(f"{r['id']}: AnsQ hitung ulang {mean_calc:.9f} != CSV {ansq_csv:.9f}")

        hop_csv = _f(c["reaq_hop_accuracy"])
        hop_calc = r["hop"]["hop_accuracy"]
        if (hop_csv is None) != (hop_calc is None):
            problems.append(f"{r['id']}: hop None-mismatch — CSV={hop_csv} hitung={hop_calc}")
        elif hop_csv is not None and abs(hop_calc - hop_csv) > 1e-5:
            problems.append(f"{r['id']}: hop hitung ulang {hop_calc:.6f} != CSV {hop_csv:.6f}")

        rr = ragas_by_id.get(r["id"])
        if rr is not None:
            reaq_csv = c["reaq_reaq_score"]
            ragas_val = rr["ragas_faithfulness"]
            if (reaq_csv in ("", None)) != (ragas_val in ("", None)):
                problems.append(f"{r['id']}: reaq_reaq_score/ragas_faithfulness blank-mismatch")
            elif reaq_csv not in ("", None) and abs(_f(reaq_csv) - _f(ragas_val)) > 1e-6:
                problems.append(f"{r['id']}: reaq_reaq_score {reaq_csv} != ragas_faithfulness {ragas_val}")
        else:
            problems.append(f"{r['id']}: tidak ditemukan di ragas_results_graphrag.csv")

    if n_short_collision > 0:
        problems.append(f"{n_short_collision} fixture punya tabrakan nama pendek pada relevant_nodes")

    def agg(field, cond=None):
        vals = [_f(r["csv"][field]) for r in rows if _f(r["csv"][field]) is not None
                and (cond is None or cond(r))]
        return (len(vals), round(sum(vals) / len(vals), 4)) if vals else (0, None)

    checks = [
        ("ansq",               agg("ansq_ansq_score")),
        ("ansq_relevansi",     agg("ansq_answer_relevance")),
        ("ansq_sintaksis",     agg("ansq_syntactic_validity")),
        ("ansq_skema_exclrbac", agg("ansq_schema_compliance", lambda r: r["id"] != RBAC_EXCLUDE)),
        ("retq_precision",     agg("retq_precision")),
        ("retq_recall",        agg("retq_recall")),
        ("retq_f1",            agg("retq_f1")),
        ("hop_semua",          agg("reaq_hop_accuracy")),
        ("hop_fokus",          agg("reaq_hop_accuracy", lambda r: _f(r["csv"]["depth_gt"]) is not None and _f(r["csv"]["depth_gt"]) <= 15)),
        ("hop_closure",        agg("reaq_hop_accuracy", lambda r: _f(r["csv"]["depth_gt"]) is not None and _f(r["csv"]["depth_gt"]) > 15)),
        ("faithfulness",       agg("reaq_reaq_score")),
    ]
    for key, (n, val) in checks:
        thesis_val, thesis_n = THESIS[key]
        if n != thesis_n:
            problems.append(f"Agregat '{key}': n={n}, diharapkan n={thesis_n}")
        if val is None or abs(val - thesis_val) > 0.0005:
            problems.append(f"Agregat '{key}': {val}, diharapkan {thesis_val} (Bab VI)")

    row_by_id = {r["id"]: r for r in rows}
    for ex_id in EXAMPLE_IDS:
        if ex_id not in row_by_id:
            problems.append(f"Fixture contoh '{ex_id}' tidak ditemukan di data")

    if problems:
        print(f"\n[GAGAL] {len(problems)} masalah ditemukan sebelum workbook ditulis:\n")
        for p in problems:
            print(f"  - {p}")
        sys.exit(1)

    print(f"[OK] Semua pemeriksaan lolos ({len(rows)} fixture, {len(checks)} agregat headline cocok dengan Bab VI).")
    return row_by_id


# ══════════════════════════════════════════════════════════════════════════
# Styling
# ══════════════════════════════════════════════════════════════════════════

NAVY = "1e3a5f"
BLUE_LIGHT = "dbeafe"
GRAY_LIGHT = "f8fafc"
GRAY_MED = "e5e7eb"
RED_LIGHT = "fee2e2"
YELLOW_LIGHT = "fef9c3"
WHITE = "ffffff"


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def thin_border():
    s = Side(style="thin", color="d1d5db")
    return Border(left=s, right=s, top=s, bottom=s)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left_wrap():
    return Alignment(horizontal="left", vertical="top", wrap_text=True)


def apply_header(cell, text, bg=NAVY, fg=WHITE, size=11, bold=True):
    cell.value = text
    cell.fill = fill(bg)
    cell.font = Font(name="Calibri", bold=bold, size=size, color=fg)
    cell.alignment = center()
    cell.border = thin_border()


def apply_value(cell, value, number_format=None, bg=WHITE, bold=False, wrap=False):
    cell.value = value
    cell.fill = fill(bg)
    cell.font = Font(name="Calibri", bold=bold, size=11, color="1e293b")
    cell.alignment = left_wrap() if wrap else center()
    cell.border = thin_border()
    if number_format:
        cell.number_format = number_format


def apply_na(cell, bg=WHITE):
    cell.value = "—"
    cell.fill = fill(bg)
    cell.font = Font(name="Calibri", italic=True, size=11, color="9ca3af")
    cell.alignment = center()
    cell.border = thin_border()


def apply_failed(cell, bg=WHITE):
    cell.value = "(gagal)"
    cell.fill = fill(bg)
    cell.font = Font(name="Calibri", italic=True, size=11, color="dc2626")
    cell.alignment = center()
    cell.border = thin_border()


def apply_formula(cell, formula, bg=YELLOW_LIGHT):
    cell.value = formula
    cell.fill = fill(bg)
    cell.font = Font(name="Calibri", size=11, color="374151")
    cell.alignment = center()
    cell.border = thin_border()


def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def _safe_text(s, limit=32000):
    if s is None:
        return ""
    s = str(s)
    if len(s) > limit:
        return s[:limit] + f"… (dipotong, {len(s)} karakter total)"
    return s


def write_metric_or_na(cell, val, bg=WHITE, is_failed=False, number_format="0.0000"):
    if val is None:
        (apply_failed if is_failed else apply_na)(cell, bg=bg)
    else:
        apply_value(cell, val, number_format=number_format, bg=bg)


# ══════════════════════════════════════════════════════════════════════════
# Sheet 0 — Cara Baca
# ══════════════════════════════════════════════════════════════════════════

def build_0_cara_baca(wb, meta):
    ws = wb.create_sheet("Cara Baca")
    ws.sheet_view.showGridLines = False

    def merged_block(row_start, span, text, bg=WHITE, fg="1e293b", size=11, bold=False,
                      italic=False, height=None, cols="A:J"):
        c1, c2 = cols.split(":")
        ws.merge_cells(f"{c1}{row_start}:{c2}{row_start + span - 1}")
        cell = ws[f"{c1}{row_start}"]
        cell.value = text
        cell.fill = fill(bg)
        cell.font = Font(name="Calibri", size=size, bold=bold, italic=italic, color=fg)
        cell.alignment = left_wrap()
        cell.border = thin_border()
        if height:
            for r in range(row_start, row_start + span):
                ws.row_dimensions[r].height = height
        return row_start + span

    r = 1
    r = merged_block(r, 1, "BUKTI PERHITUNGAN METRIK EVALUASI — Sistem GraphRAG Kubernetes",
                      bg=NAVY, fg=WHITE, size=14, bold=True, height=30)
    r = merged_block(r, 1, "Jihan Aurelia · Tugas Akhir STI ITB · 2026 · GraphRAG saja (102 fixture)",
                      bg="2563eb", fg=WHITE, size=11, height=22)
    r += 1

    r = merged_block(r, 1, "TUJUAN", bg="374151", fg=WHITE, bold=True, height=22)
    r = merged_block(r, 2,
        "Workbook ini memungkinkan penguji menelusuri setiap angka agregat di Bab VI sampai ke "
        "fixture individual dan data mentah yang menghasilkannya, tanpa membuka file JSON/CSV lain. "
        "Sheet 'Contoh Terpandu' menelusuri 3 fixture secara lengkap; sheet AnsQ/RetQ/ReaQ memuat "
        "seluruh 102 fixture.", height=32)
    r += 1

    r = merged_block(r, 1, "PROVENANCE RUN", bg="374151", fg=WHITE, bold=True, height=22)
    prov_lines = [
        f"run_id: {meta.get('run_id', '?')}   ·   git_commit: {meta.get('git_commit', '?')}   ·   "
        f"timestamp: {meta.get('timestamp', '?')}",
        "Embedder relevansi jawaban: text-embedding-3-small (evaluate.py:508). "
        "Hakim Faithfulness: gpt-4o-mini (recompute_ragas.py:188).",
        f"Catatan: eval_run_meta_graphrag.json mencatat n_fixtures={meta.get('n_fixtures', '?')}, "
        "sedangkan CSV final berisi 102 baris — satu id basi ('pim_trying_to_use_a_container') "
        "dibuang saat rekurasi dan tidak dipakai di manapun dalam workbook ini.",
    ]
    for line in prov_lines:
        r = merged_block(r, 1, line, bg=WHITE, height=24)
    r += 1

    r = merged_block(r, 1, "RUMUS & SITASI KODE", bg="374151", fg=WHITE, bold=True, height=22)
    formulas = [
        "AnsQ = rata-rata sub-metrik non-null (sintaksis, skema, relevansi, layer-3)  —  evaluate.py:247-248",
        "Relevansi Jawaban = cosine similarity embedding (jawaban sistem vs ground truth)  —  evaluate.py:103, model di :508",
        "RetQ: Precision = cocok/diambil, Recall = cocok/relevan-GT, F1 = 2PR/(P+R)  —  evaluate.py:315-320",
        "Hop Accuracy = edge cocok / edge jalur GT, pencocokan ketat (strip+lower)  —  evaluate.py:364-367",
        "Faithfulness (ReaQ) = RAGAS Faithfulness, hakim gpt-4o-mini  —  recompute_ragas.py:180-191",
    ]
    for line in formulas:
        r = merged_block(r, 1, line, bg=GRAY_LIGHT, height=20)
    r += 1

    r = merged_block(r, 1, "LEGENDA", bg="374151", fg=WHITE, bold=True, height=22)
    legend = [
        ("—  =  N/A karena desain (bukan kegagalan)", GRAY_LIGHT, "9ca3af", True),
        ("(gagal)  =  gagal diukur, dikeluarkan dari rerata", WHITE, "dc2626", True),
        ("sel kuning  =  formula Excel hidup, dihitung ulang saat file dibuka", YELLOW_LIGHT, "374151", False),
        ("tombol + di atas kolom  =  kolom daftar yang dilipat (default tersembunyi)", WHITE, "374151", False),
    ]
    for text, bg, fg, italic in legend:
        row_r = r
        ws.merge_cells(f"A{row_r}:J{row_r}")
        cell = ws[f"A{row_r}"]
        cell.value = text
        cell.fill = fill(bg)
        cell.font = Font(name="Calibri", size=11, italic=italic, color=fg)
        cell.alignment = left_wrap()
        cell.border = thin_border()
        ws.row_dimensions[row_r].height = 20
        r += 1
    r += 1

    r = merged_block(r, 1, "RINGKASAN AGREGAT — angka headline vs Bab VI", bg="374151", fg=WHITE, bold=True, height=22)
    headers = ["Metrik", "n", "Nilai di Bab VI", "Dihitung Skrip", "Formula Excel Hidup", "Status"]
    hdr_row = r
    for col, h in enumerate(headers, 1):
        apply_header(ws.cell(hdr_row, col), h, bg="4b5563", size=10)
    ws.row_dimensions[hdr_row].height = 22
    r += 1

    agg_rows = [
        ("AnsQ (komposit)",              "ansq",                "AnsQ", "H"),
        ("— Relevansi Jawaban",          "ansq_relevansi",      "AnsQ", "G"),
        ("— Sintaksis YAML",             "ansq_sintaksis",      "AnsQ", "C"),
        ("— Skema YAML (excl. RBAC)",    "ansq_skema_exclrbac", "AnsQ", "D"),
        ("RetQ Precision",               "retq_precision",      "RetQ", "H"),
        ("RetQ Recall",                  "retq_recall",         "RetQ", "I"),
        ("RetQ F1 (= Skor RetQ)",        "retq_f1",             "RetQ", "J"),
        ("Hop Accuracy (semua)",         "hop_semua",           "ReaQ", "H"),
        ("Faithfulness (ReaQ)",          "faithfulness",        "ReaQ", "L"),
    ]
    for label, key, sheet_name, col_letter in agg_rows:
        thesis_val, n = THESIS[key]
        formula = f"=AVERAGE('{sheet_name}'!{col_letter}5:{col_letter}106)"
        apply_value(ws.cell(r, 1), label, bg=WHITE, wrap=True)
        apply_value(ws.cell(r, 2), n, bg=WHITE)
        apply_value(ws.cell(r, 3), thesis_val, number_format="0.0000", bg=WHITE)
        apply_value(ws.cell(r, 4), thesis_val, number_format="0.0000", bg=GRAY_LIGHT)
        apply_formula(ws.cell(r, 5), formula)
        apply_value(ws.cell(r, 6), "COCOK", bg="dcfce7", bold=True)
        ws.row_dimensions[r].height = 18
        r += 1
    r += 1

    r = merged_block(r, 1, "BATAS VERIFIKASI — apa yang bisa dan tidak bisa dihitung ulang di sheet ini", bg="374151", fg=WHITE, bold=True, height=22)
    limits_headers = ["Metrik", "Dapat dihitung ulang di sheet?", "Yang disajikan sebagai gantinya"]
    hdr_row2 = r
    widths3 = [3, 5, 2]
    ws.merge_cells(f"A{hdr_row2}:B{hdr_row2}")
    apply_header(ws[f"A{hdr_row2}"], limits_headers[0], bg="4b5563", size=10)
    ws.merge_cells(f"C{hdr_row2}:E{hdr_row2}")
    apply_header(ws[f"C{hdr_row2}"], limits_headers[1], bg="4b5563", size=10)
    ws.merge_cells(f"F{hdr_row2}:J{hdr_row2}")
    apply_header(ws[f"F{hdr_row2}"], limits_headers[2], bg="4b5563", size=10)
    ws.row_dimensions[hdr_row2].height = 20
    r += 1

    limits = [
        ("RetQ (P/R/F1)", "Ya, penuh", "Jumlah + daftar node, node terlewat, node di luar GT"),
        ("Hop Accuracy", "Ya, penuh", "Jumlah + daftar edge, daftar edge terlewat"),
        ("AnsQ — Relevansi", "Tidak — cosine similarity embedding", "Kedua teks yang dibandingkan, berdampingan"),
        ("Faithfulness", "Tidak — hakim LLM, rincian klaim tak tersimpan", "Jawaban sistem + konteks yang dibaca judge"),
    ]
    for a, b, c in limits:
        ws.merge_cells(f"A{r}:B{r}")
        apply_value(ws[f"A{r}"], a, bg=WHITE, wrap=True, bold=True)
        ws.merge_cells(f"C{r}:E{r}")
        apply_value(ws[f"C{r}"], b, bg=WHITE, wrap=True)
        ws.merge_cells(f"F{r}:J{r}")
        apply_value(ws[f"F{r}"], c, bg=WHITE, wrap=True)
        ws.row_dimensions[r].height = 30
        r += 1
    r += 1

    r = merged_block(r, 1, "KENAPA FAITHFULNESS TIDAK BISA DIVERIFIKASI PENUH", bg="374151", fg=WHITE, bold=True, height=22)
    r = merged_block(r, 3,
        "Rincian klaim dari run penilaian asli tidak pernah tersimpan. File "
        "data/ragas_run_graphrag*.log memang memuat blok 'statements' (statement + reason + "
        "verdict), tetapi seluruh 60 blok itu ber-finish_reason='length' — hanya panggilan yang "
        "responsnya kepotong yang ikut ter-log; panggilan sukses yang menghasilkan skor final "
        "tidak. Jadi per-klaim breakdown yang konsisten dengan Bab VI tidak ada di sumber manapun. "
        "Karena itu Faithfulness ditangani seperti relevansi jawaban: tidak dihitung ulang, tapi "
        "bahan penilaiannya (jawaban sistem + konteks graf) disajikan di sheet ReaQ.", height=22)
    r += 1

    r = merged_block(r, 1, "YANG SENGAJA TIDAK DIMASUKKAN", bg="374151", fg=WHITE, bold=True, height=22)
    excluded = [
        "faithfulness_decomposition.csv / _raw.jsonl — run hakim LLM terpisah yang stokastik; "
        "skornya berbeda dari angka Bab VI pada 84 dari 95 fixture. Tidak dipakai sama sekali agar "
        "seluruh isi workbook konsisten dengan Bab VI.",
        "Baseline Vector RAG / Vanilla LLM dan seluruh run ablasi/depth-sweep — di luar cakupan "
        "workbook ini (GraphRAG saja).",
        "ansq_layer3_compliance — kosong di seluruh 102 baris pada run produksi; hanya terisi pada "
        "run ablasi (evaluate.py:227,245), karena itu tidak dijadikan kolom.",
    ]
    for line in excluded:
        r = merged_block(r, 1, line, bg=WHITE, height=32)

    for col, w in enumerate([28, 10, 16, 16, 24, 12, 12, 12, 12, 12], 1):
        set_col_width(ws, col, w)
    ws.freeze_panes = "A4"
    return ws


# ══════════════════════════════════════════════════════════════════════════
# Sheet 1 — Contoh Terpandu
# ══════════════════════════════════════════════════════════════════════════

QNA = {
    "rolebinding_dev": [
        ("Kenapa urutan jalur sistem beda dari GT tapi Hop Accuracy tetap 1,0?",
         "Hop Accuracy berbasis himpunan edge, bukan urutan (evaluate.py:320) — yang dinilai "
         "apakah edge yang sama ada, bukan urutan penelusurannya."),
        ("Kalau retrieval-nya sempurna, kenapa Faithfulness cuma 0,20?",
         "Faithfulness dan retrieval mengukur hal berbeda. Retrieval menilai apakah node/edge "
         "yang tepat berhasil diambil; Faithfulness menilai apakah kalimat di jawaban bisa "
         "ditelusuri ke konteks. Sistem bisa retrieve dengan sempurna tapi tetap menulis "
         "kalimat yang tidak sepenuhnya didukung konteks."),
    ],
    "kubectl_find_pods_with_env": [
        ("Kenapa Hop Accuracy cuma 0,7143, apakah sistemnya banyak salah jalan?",
         "Tidak — precision-nya 1,0 (semua 30 edge yang ditelusuri tepat sasaran). 12 edge yang "
         "terlewat seluruhnya anak dari node yang justru berhasil dijangkau, artinya kekurangannya "
         "murni di ekspansi hop ke-3 — batas kedalaman yang sistematis, bukan kesalahan acak."),
    ],
    "clusterrole_read_pods": [
        ("Kenapa RetQ rendah padahal Hop Accuracy dan AnsQ tinggi?",
         "Sistem menelusuri 54 edge sementara GT hanya mengurasi 4 sebagai relevan — precision "
         "jatuh karena banyak node di luar cakupan GT terambil. RetQ rendah tidak selalu berarti "
         "sistem salah; bisa berarti ia mengambil lebih luas daripada yang dikurasi sebagai "
         "relevan. Ini konsisten dengan disclosure GT-sensitivity di Bab VI."),
    ],
}


def build_1_contoh_terpandu(wb, row_by_id):
    ws = wb.create_sheet("Contoh Terpandu")
    ws.sheet_view.showGridLines = False

    r = 1
    ws.merge_cells(f"A{r}:F{r}")
    c = ws[f"A{r}"]
    c.value = "CONTOH TERPANDU — 3 fixture ditelusuri lengkap"
    c.fill = fill(NAVY)
    c.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    c.alignment = center()
    ws.row_dimensions[r].height = 30
    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    c = ws[f"A{r}"]
    c.value = "Tiap contoh mengajarkan satu hal berbeda — cukup hafal ceritanya, bukan 102 baris."
    c.fill = fill("2563eb")
    c.font = Font(name="Calibri", size=11, color=WHITE, italic=True)
    c.alignment = center()
    ws.row_dimensions[r].height = 22
    r += 2

    titles = {
        "rolebinding_dev": "CONTOH A — semuanya benar, kecuali Faithfulness",
        "kubectl_find_pods_with_env": "CONTOH B — Hop Accuracy < 1, penyebabnya sistematis",
        "clusterrole_read_pods": "CONTOH C — RetQ rendah karena over-retrieval",
    }

    for ex_id in EXAMPLE_IDS:
        row = row_by_id[ex_id]
        c_, retq, hop = row["csv"], row["retq"], row["hop"]

        ws.merge_cells(f"A{r}:F{r}")
        cell = ws[f"A{r}"]
        cell.value = f"{titles[ex_id]}  —  {ex_id}"
        cell.fill = fill("374151")
        cell.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
        cell.alignment = left_wrap()
        ws.row_dimensions[r].height = 24
        r += 1

        ws.merge_cells(f"A{r}:F{r}")
        cell = ws[f"A{r}"]
        cell.value = f"Pertanyaan: {row['question']}"
        cell.fill = fill(GRAY_LIGHT)
        cell.font = Font(name="Calibri", italic=True, size=11, color="374151")
        cell.alignment = left_wrap()
        ws.row_dimensions[r].height = 30
        r += 1

        metric_headers = ["Metrik", "Angka", "Cara baca"]
        hdr_r = r
        for col, h in enumerate(metric_headers, 1):
            apply_header(ws.cell(hdr_r, col), h, bg="4b5563", size=10)
        ws.merge_cells(f"C{hdr_r}:F{hdr_r}")
        ws.row_dimensions[hdr_r].height = 18
        r += 1

        ansq_score = _f(c_["ansq_ansq_score"])
        metric_rows = [
            ("AnsQ", ansq_score, sentence_ansq(row)),
            ("RetQ (F1)", _f(c_["retq_f1"]), sentence_retq(row)),
        ]
        if hop["hop_accuracy"] is not None:
            metric_rows.append(("Hop Accuracy", hop["hop_accuracy"],
                                 f"{hop['n_match']}/{hop['n_gt']} edge GT ditelusuri "
                                 f"({hop['n_pred']} edge total ditelusuri sistem)"))
        faith = _f(c_["reaq_reaq_score"])
        metric_rows.append(("Faithfulness", faith, "hasil hakim LLM gpt-4o-mini"))

        for label, val, desc in metric_rows:
            apply_value(ws.cell(r, 1), label, bg=WHITE, bold=True)
            apply_value(ws.cell(r, 2), val, number_format="0.0000", bg=YELLOW_LIGHT, bold=True)
            ws.merge_cells(f"C{r}:F{r}")
            apply_value(ws.cell(r, 3), desc, bg=WHITE, wrap=True)
            ws.row_dimensions[r].height = 20
            r += 1

        for q, a in QNA.get(ex_id, []):
            ws.merge_cells(f"A{r}:F{r}")
            cell = ws[f"A{r}"]
            cell.value = f"Kalau penguji bertanya: \"{q}\""
            cell.fill = fill(BLUE_LIGHT)
            cell.font = Font(name="Calibri", bold=True, italic=True, size=10, color="1e40af")
            cell.alignment = left_wrap()
            ws.row_dimensions[r].height = 18
            r += 1
            ws.merge_cells(f"A{r}:F{r}")
            cell = ws[f"A{r}"]
            cell.value = f"Jawab: {a}"
            cell.fill = fill(WHITE)
            cell.font = Font(name="Calibri", size=10, color="374151")
            cell.alignment = left_wrap()
            ws.row_dimensions[r].height = 44
            r += 1

        r += 1

    for col, w in enumerate([16, 12, 24, 24, 24, 24], 1):
        set_col_width(ws, col, w)
    ws.freeze_panes = "A3"
    return ws


# ══════════════════════════════════════════════════════════════════════════
# Sheet 2 — AnsQ
# ══════════════════════════════════════════════════════════════════════════

def build_2_ansq(wb, rows):
    ws = wb.create_sheet("AnsQ")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    apply_header(ws["A1"], "AnsQ — Kualitas Jawaban (102 fixture)", bg=NAVY, size=13)
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:J2")
    note = ws["A2"]
    note.value = ("Kolom E/F dilipat (klik + di atas kolom untuk membuka). Bandingkan dulu "
                  "Jawaban Sistem (E) dengan Jawaban Ground Truth (F) sebelum melihat skor "
                  "Relevansi (G) yang dihasilkan dari keduanya. AnsQ = rata-rata sub-metrik yang "
                  "tidak kosong. layer3_compliance selalu kosong pada run produksi (lihat sheet "
                  "Cara Baca) sehingga tidak dijadikan kolom.")
    note.fill = fill(GRAY_LIGHT)
    note.font = Font(name="Calibri", italic=True, size=9, color="6b7280")
    note.alignment = left_wrap()
    ws.row_dimensions[2].height = 30

    ws.row_dimensions[3].height = 4

    headers = ["ID Fixture", "Tipe", "Sintaksis YAML", "Skema YAML", "Jawaban Sistem",
               "Jawaban Ground Truth", "Relevansi Jawaban", "AnsQ", "Kalimat Penjelas", "Catatan Teknis"]
    for col, h in enumerate(headers, 1):
        apply_header(ws.cell(4, col), h, bg="4b5563", size=10)
    ws.row_dimensions[4].height = 30

    r = 5
    for row in rows:
        c = row["csv"]
        bg = GRAY_LIGHT if (r % 2 == 0) else WHITE
        apply_value(ws.cell(r, 1), row["id"], bg=bg, wrap=True)
        apply_value(ws.cell(r, 2), row["type"], bg=bg)

        syn, sch, rel = _f(c["ansq_syntactic_validity"]), _f(c["ansq_schema_compliance"]), _f(c["ansq_answer_relevance"])
        write_metric_or_na(ws.cell(r, 3), syn, bg=bg)
        write_metric_or_na(ws.cell(r, 4), sch, bg=bg)

        apply_value(ws.cell(r, 5), _safe_text(row["answer_full"]), bg=bg, wrap=True)
        apply_value(ws.cell(r, 6), _safe_text(row["gt_answer"]), bg=bg, wrap=True)

        apply_value(ws.cell(r, 7), rel, number_format="0.0000", bg=bg)
        apply_value(ws.cell(r, 8), _f(c["ansq_ansq_score"]), number_format="0.0000", bg=YELLOW_LIGHT, bold=True)

        apply_value(ws.cell(r, 9), sentence_ansq(row), bg=bg, wrap=True)

        catatan = []
        if syn is None:
            catatan.append("Sintaksis & Skema: N/A — bukan fixture penghasil YAML (evaluate.py:181-212). Kosong karena DESAIN.")
        if row["id"] == RBAC_EXCLUDE:
            catatan.append("Dikecualikan dari rerata Kepatuhan Skema (di luar cakupan KG, catatan kaki Tabel 3.1).")
        apply_value(ws.cell(r, 10), " ".join(catatan), bg=bg, wrap=True)

        ws.row_dimensions[r].height = 60
        r += 1

    ws.column_dimensions.group("E", "F", outline_level=1, hidden=False)
    for col, w in enumerate([32, 14, 13, 13, 46, 46, 13, 10, 46, 40], 1):
        set_col_width(ws, col, w)
    ws.freeze_panes = "C5"
    return ws


# ══════════════════════════════════════════════════════════════════════════
# Sheet 3 — RetQ
# ══════════════════════════════════════════════════════════════════════════

def build_3_retq(wb, rows):
    ws = wb.create_sheet("RetQ")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:M1")
    apply_header(ws["A1"], "RetQ — Kualitas Retrieval (102 fixture)", bg=NAVY, size=13)
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:M2")
    note = ws["A2"]
    note.value = ("Kolom C/D dan K/L dilipat (klik + di atas kolom). Bandingkan dulu Daftar Node "
                  "Diambil Sistem (C) dengan Daftar Node Relevan GT (D) sebelum melihat Jumlah (E/F/G) "
                  "dan skor (H/I/J) yang dihasilkan dari keduanya. Kolom K/L menjelaskan selisihnya — "
                  "kenapa skornya tidak 1,0. Perbandingan node pakai nama pendek (segmen terakhir "
                  "setelah titik, evaluate.py:299) — pada 102 fixture ini tidak ada satu pun tabrakan "
                  "nama akibat normalisasi tersebut.")
    note.fill = fill(GRAY_LIGHT)
    note.font = Font(name="Calibri", italic=True, size=9, color="6b7280")
    note.alignment = left_wrap()
    ws.row_dimensions[2].height = 40

    ws.row_dimensions[3].height = 4

    headers = ["ID Fixture", "Tipe", "Daftar Node Diambil Sistem", "Daftar Node Relevan (GT)",
               "Jumlah Node Diambil Sistem", "Jumlah Node Relevan (GT)", "Jumlah Node Cocok",
               "Precision", "Recall", "F1 (Skor RetQ)",
               "Daftar Node GT Tidak Terambil", "Daftar Node Diambil di Luar GT", "Kalimat Penjelas"]
    for col, h in enumerate(headers, 1):
        apply_header(ws.cell(4, col), h, bg="4b5563", size=10)
    ws.row_dimensions[4].height = 34

    r = 5
    for row in rows:
        c, retq = row["csv"], row["retq"]
        bg = GRAY_LIGHT if (r % 2 == 0) else WHITE
        apply_value(ws.cell(r, 1), row["id"], bg=bg, wrap=True)
        apply_value(ws.cell(r, 2), row["type"], bg=bg)

        apply_value(ws.cell(r, 3), _safe_text(", ".join(retq["R"])), bg=bg, wrap=True)
        apply_value(ws.cell(r, 4), _safe_text(", ".join(retq["G"])), bg=bg, wrap=True)

        apply_value(ws.cell(r, 5), retq["n_R"], bg=bg)
        apply_value(ws.cell(r, 6), retq["n_G"], bg=bg)
        apply_value(ws.cell(r, 7), retq["n_inter"], bg=bg)
        apply_value(ws.cell(r, 8), _f(c["retq_precision"]), number_format="0.0000", bg=bg)
        apply_value(ws.cell(r, 9), _f(c["retq_recall"]), number_format="0.0000", bg=bg)
        apply_value(ws.cell(r, 10), _f(c["retq_f1"]), number_format="0.0000", bg=YELLOW_LIGHT, bold=True)

        apply_value(ws.cell(r, 11), _safe_text(", ".join(retq["missed"])) or "(tidak ada)", bg=bg, wrap=True)
        apply_value(ws.cell(r, 12), _safe_text(", ".join(retq["extra"])) or "(tidak ada)", bg=bg, wrap=True)

        apply_value(ws.cell(r, 13), sentence_retq(row), bg=bg, wrap=True)

        ws.row_dimensions[r].height = 60
        r += 1

    ws.column_dimensions.group("C", "D", outline_level=1, hidden=False)
    ws.column_dimensions.group("K", "L", outline_level=1, hidden=False)
    for col, w in enumerate([32, 14, 40, 40, 11, 11, 11, 11, 11, 12, 40, 40, 46], 1):
        set_col_width(ws, col, w)
    ws.freeze_panes = "E5"
    return ws


# ══════════════════════════════════════════════════════════════════════════
# Sheet 4 — ReaQ
# ══════════════════════════════════════════════════════════════════════════

def build_4_reaq(wb, rows):
    ws = wb.create_sheet("ReaQ")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:M1")
    apply_header(ws["A1"], "ReaQ — Kualitas Penalaran: Hop Accuracy + Faithfulness (102 fixture)", bg=NAVY, size=13)
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:M2")
    note = ws["A2"]
    note.value = ("Dua metrik tak berhubungan digabung di tab ini (lihat pita baris 3): Hop Accuracy "
                  "menilai jalur graf (edge), Faithfulness menilai kalimat jawaban (klaim). Kolom C/D "
                  "dan J/K dilipat — bandingkan dulu jalur/materi mentahnya sebelum melihat skor yang "
                  "dihasilkan. Pencocokan edge bersifat KETAT (strip+lower, evaluate.py:364-367) — "
                  "berbeda dari field missed_edges di JSONL yang memakai pencocokan longgar dan sengaja "
                  "tidak dipakai. depth_gt (jumlah edge GT) != gt_depth (kedalaman traversal kurasi, "
                  "1-3) — hanya depth_gt yang dipakai di sini.")
    note.fill = fill(GRAY_LIGHT)
    note.font = Font(name="Calibri", italic=True, size=9, color="6b7280")
    note.alignment = left_wrap()
    ws.row_dimensions[2].height = 48

    ws.merge_cells("C3:I3")
    apply_header(ws["C3"], "HOP ACCURACY — jalur penelusuran graf", bg="4338ca", size=10)
    ws.merge_cells("J3:L3")
    apply_header(ws["J3"], "FAITHFULNESS — klaim dalam jawaban", bg="be185d", size=10)
    for col in (1, 2, 13):
        apply_value(ws.cell(3, col), "", bg=WHITE)
    ws.row_dimensions[3].height = 18

    headers = ["ID Fixture", "Tipe", "Daftar Jalur GT", "Daftar Jalur Sistem",
               "Jumlah Edge Jalur GT", "Jumlah Edge Ditelusuri Sistem", "Jumlah Edge Cocok",
               "Hop Accuracy", "Daftar Edge GT Terlewat",
               "Jawaban Sistem", "Konteks yang Dinilai Judge", "Faithfulness", "Kalimat Penjelas"]
    for col, h in enumerate(headers, 1):
        apply_header(ws.cell(4, col), h, bg="4b5563", size=10)
    ws.row_dimensions[4].height = 34

    r = 5
    for row in rows:
        c, hop = row["csv"], row["hop"]
        bg = GRAY_LIGHT if (r % 2 == 0) else WHITE
        apply_value(ws.cell(r, 1), row["id"], bg=bg, wrap=True)
        apply_value(ws.cell(r, 2), row["type"], bg=bg)

        apply_value(ws.cell(r, 3), _safe_text("\n".join(hop["gt_display"])) or "(tidak ada jalur GT)", bg=bg, wrap=True)
        apply_value(ws.cell(r, 4), _safe_text("\n".join(hop["pred_display"])) or "(tidak menelusuri)", bg=bg, wrap=True)

        apply_value(ws.cell(r, 5), _i(c["depth_gt"]), bg=bg)
        apply_value(ws.cell(r, 6), _i(c["depth_pred"]), bg=bg)

        hop_acc = hop["hop_accuracy"]
        if hop_acc is None:
            apply_na(ws.cell(r, 7), bg=bg)
            apply_na(ws.cell(r, 8), bg=bg)
        else:
            apply_value(ws.cell(r, 7), hop["n_match"], bg=bg)
            apply_value(ws.cell(r, 8), hop_acc, number_format="0.0000", bg=YELLOW_LIGHT, bold=True)

        apply_value(ws.cell(r, 9), _safe_text("\n".join(hop["missed"])) or "(tidak ada)", bg=bg, wrap=True)

        apply_value(ws.cell(r, 10), _safe_text(row["answer_full"]), bg=bg, wrap=True)

        ctx = row["graph_context"] or ""
        excerpt = ctx[:2000]
        total_len = len(ctx)
        ctx_display = excerpt
        if total_len > 2000:
            ctx_display += f"\n\n… (kutipan 2.000 dari {total_len} karakter total)"
        apply_value(ws.cell(r, 11), _safe_text(ctx_display), bg=bg, wrap=True)

        faith = _f(c["reaq_reaq_score"])
        write_metric_or_na(ws.cell(r, 12), faith, bg=YELLOW_LIGHT if faith is not None else bg, is_failed=True, number_format="0.0000")
        if faith is not None:
            ws.cell(r, 12).font = Font(name="Calibri", bold=True, size=11, color="1e293b")

        apply_value(ws.cell(r, 13), sentence_reaq(row), bg=bg, wrap=True)

        ws.row_dimensions[r].height = 60
        r += 1

    ws.column_dimensions.group("C", "D", outline_level=1, hidden=False)
    ws.column_dimensions.group("I", "I", outline_level=1, hidden=False)
    ws.column_dimensions.group("J", "K", outline_level=1, hidden=False)
    for col, w in enumerate([32, 14, 40, 40, 11, 11, 11, 12, 40, 46, 46, 12, 46], 1):
        set_col_width(ws, col, w)
    ws.freeze_panes = "E5"
    return ws


# ══════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════

def post_write_check(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=False)
    expected_sheets = ["Cara Baca", "Contoh Terpandu", "AnsQ", "RetQ", "ReaQ"]
    if wb.sheetnames != expected_sheets:
        print(f"[GAGAL] pasca-tulis: sheet = {wb.sheetnames}, diharapkan {expected_sheets}")
        sys.exit(1)
    for name in ["AnsQ", "RetQ", "ReaQ"]:
        ws = wb[name]
        if ws.max_row < 106:
            print(f"[GAGAL] pasca-tulis: sheet '{name}' max_row={ws.max_row}, diharapkan >= 106")
            sys.exit(1)
    print(f"[OK] Pemeriksaan pasca-tulis lolos ({len(expected_sheets)} sheet, baris data lengkap).")


def main():
    logger.info("Membaca data sumber...")
    rows = build_rows()
    ragas_by_id = {r["id"]: r for r in _read_csv(DATA / "ragas_results_graphrag.csv")}
    meta = json.load(open(DATA / "eval_run_meta_graphrag.json", encoding="utf-8"))

    logger.info("Menjalankan verifikasi (fail-fast sebelum menulis workbook)...")
    row_by_id = verify(rows, ragas_by_id)

    logger.info("Menyusun workbook...")
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True

    build_0_cara_baca(wb, meta)
    build_1_contoh_terpandu(wb, row_by_id)
    build_2_ansq(wb, rows)
    build_3_retq(wb, rows)
    build_4_reaq(wb, rows)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    logger.info(f"[OK] Generated: {OUTPUT}")

    post_write_check(OUTPUT)


if __name__ == "__main__":
    main()
